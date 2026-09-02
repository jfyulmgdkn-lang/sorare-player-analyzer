import os
import time
import asyncio
from collections import defaultdict
from datetime import datetime, timezone
from typing import Dict, List, Optional, Set, Tuple

import requests
import discord
from discord import app_commands
from discord.ext import commands
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

API_URL = "https://api.sorare.com/graphql"
CURRENT_SEASON_YEAR = 2026
SEASON_START = "2026-07-01T00:00:00Z"
SEASON_END = "2027-06-30T23:59:59Z"

REQUEST_TIMEOUT = 30
REQUEST_INTERVAL = 0.20

L40_WEIGHT = 0.70
START_RATE_WEIGHT = 0.30

LEAGUES = {
    "1": ("Bundesliga", "bundesliga-de"),
    "2": ("2. Bundesliga", "2-bundesliga"),
    "3": ("Premier League", "premier-league-gb-eng"),
    "4": ("La Liga", "laliga-es"),
    "5": ("Ligue 1", "ligue-1-fr"),
    "6": ("Ligue 2", "ligue-2-fr"),
    "7": ("MLS", "mlspa"),
    "8": ("Bundesliga Österreich", "austrian-bundesliga"),
    "9": ("HNL", "1-hnl"),
    "10": ("Primeira Liga", "primeira-liga-pt"),
    "11": ("Jupiler Pro League", "jupiler-pro-league"),
}

LEAGUE_THRESHOLDS = {
    "bundesliga-de": [320, 360, 380, 420, 440, 470],
    "2-bundesliga": [320, 360, 380, 420, 440, 470],
    "premier-league-gb-eng": [320, 360, 380, 400, 430, 450],
    "laliga-es": [320, 360, 380, 420, 440, 470],
    "ligue-1-fr": [320, 360, 380, 410, 440, 460],
    "ligue-2-fr": [320, 360, 380, 420, 440, 470],
    "mlspa": [340, 380, 400, 420, 460],
    "austrian-bundesliga": [320, 360, 380, 420, 440, 470],
    "1-hnl": [320, 360, 380, 420, 440, 470],
    "primeira-liga-pt": [320, 360, 380, 410, 440, 460],
    "jupiler-pro-league": [320, 360, 380, 410, 440, 460],
}

# Sorare 27 Contender (Limited & Rare):
# Diese vier realen Ligen werden im Analyzer als EIN gemeinsamer Pool behandelt.
CONTENDER_SLUG = "contender"
# Für Ligue 2 verwenden wir bewusst nur die Vereine, die auf der
# aktuellen Sorare-Collection-Seite 2026/27 als Karten-Collections
# angezeigt werden.
LIGUE_2_SORARE_2026_27_CLUBS = {
    "AS Saint-Étienne",
    "AS Saint-Etienne",
    "FC Nantes",
    "Grenoble Foot 38",
    "Pau FC",
    "Stade de Reims",
}

CONTENDER_COMPETITIONS = [
    "2-bundesliga",
    "ligue-2-fr",
    "austrian-bundesliga",
    "1-hnl",
]

# Alle vier bisherigen Einzel-Ligen hatten bei uns dieselben A-Marken.
# Deshalb gelten diese Marken auch für die gemeinsame Contender-Auswertung.
LEAGUE_THRESHOLDS[CONTENDER_SLUG] = [320, 360, 380, 420, 440, 470]



# ============================================================
# Discord-User-ID -> Sorare-Account
# ============================================================
DISCORD_SORARE_ACCOUNTS = {
    652050886985777189: "adixyz",
    486198451626049537: "bartholomaus",
    628316534539812865: "pidel",
    463614748366340096: "sweggaausbrazil",
    235803999167709195: "golden_goal-699e9a95-45ec-4764-8771-a26226e5d4e9",
    344185350161170442: "orsa-b186064b-a1fd-427c-b15a-8f47077f11ef",
    228944337793318912: "rabise",
    246679363502735370: "lublol",
    1083049525830242384: "ksc_jockel",
    209387327229919233: "once",
    593253225973547019: "lost_drian",
    898242220740718593: "podrickson",
    1162305763306385441: "salapele-99",
    712804219698282497: "kurve1887",
    543042637087637504: "max-ntl-f9ab356a-f9e9-4a11-b76f-2a7fd50bc316",
    376728638734860290: "fidelitas-14c9ae75-c292-474e-8608-1cc1b62d11ba",
    344541166147993601: "cano35",
}

POSITION_ORDER = ["Goalkeeper", "Defender", "Midfielder", "Forward"]
POSITION_LABELS = {
    "Goalkeeper": "TW",
    "Defender": "VER",
    "Midfielder": "MID",
    "Forward": "FW",
}


# ============================================================
# Manuelle Transfer-Korrekturen
# ============================================================
# Sorare kann bei ganz frischen Transfers noch den alten activeClub liefern.
# Diese Spieler werden deshalb auf ihren aktuellen Bundesliga-Verein gesetzt.
#
# Der Zielverein wird NICHT per Sorare-Slug hartcodiert, sondern anhand des
# Vereinsnamens aus den aktuell geladenen Competition-Teams gesucht.
# So bleiben die League-Team-Slugs sauber aus der API.

PLAYER_CLUB_OVERRIDES = {
    "Ørjan Nyland": "RB Leipzig",
    "Orjan Nyland": "RB Leipzig",
    "Facundo Medina": "Bayer 04 Leverkusen",
}


COLOR_HEADER = "E7A4C7"
COLOR_DARK = "242424"
COLOR_TOTAL = "3A352D"
COLOR_WHITE = "FFFFFF"
COLOR_TEXT = "F2F2F2"
COLOR_GOLD = "D8C38A"
COLOR_BORDER = "686868"

THIN_BORDER = Side(style="thin", color=COLOR_BORDER)
GOLD_BORDER = Side(style="thin", color=COLOR_GOLD)

load_dotenv()
SORARE_API_KEY = os.getenv("SORARE_API_KEY")
DISCORD_TOKEN = os.getenv("DISCORD_TOKEN")
GUILD_ID_RAW = os.getenv("GUILD_ID")

if not SORARE_API_KEY:
    raise RuntimeError("SORARE_API_KEY fehlt in deiner .env-Datei.")

if not DISCORD_TOKEN:
    raise RuntimeError("DISCORD_TOKEN fehlt in deiner .env-Datei.")

GUILD_ID = int(GUILD_ID_RAW) if GUILD_ID_RAW else None

HEADERS = {"APIKEY": SORARE_API_KEY, "Content-Type": "application/json"}


def graphql_request(query: str, variables: Optional[dict] = None) -> dict:
    response = requests.post(
        API_URL,
        headers=HEADERS,
        json={"query": query, "variables": variables or {}},
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("errors"):
        messages = "\n".join(
            error.get("message", str(error)) for error in payload["errors"]
        )
        raise RuntimeError(f"Sorare GraphQL Fehler:\n{messages}")
    time.sleep(REQUEST_INTERVAL)
    return payload["data"]



QUERY_USER_CARDS = """
query UserCards($userSlug: String!, $after: String) {
  user(slug: $userSlug) {
    slug
    nickname
    cards(first: 100, after: $after) {
      nodes {
        slug
        rarityTyped
        seasonYear
        anyPlayer {
          slug
          displayName
        }
        anyPositions
      }
      pageInfo {
        hasNextPage
        endCursor
      }
    }
  }
}
"""


def get_user_inseason_limited_cards(user_slug: str) -> List[dict]:
    """
    Lädt ALLE Limited-Karten des Users aus Saison 2026/27.
    Pagination läuft über alle Seiten.
    """
    after = None
    result = []

    while True:
        data = graphql_request(
            QUERY_USER_CARDS,
            {
                "userSlug": user_slug,
                "after": after,
            },
        )

        user = data.get("user")
        if not user:
            raise RuntimeError(
                f"Sorare-Account '{user_slug}' wurde nicht gefunden."
            )

        cards = user.get("cards") or {}

        for card in cards.get("nodes") or []:
            if str(card.get("rarityTyped") or "").lower() != "limited":
                continue

            if card.get("seasonYear") != CURRENT_SEASON_YEAR:
                continue

            player = card.get("anyPlayer") or {}
            player_slug = player.get("slug")

            if not player_slug:
                continue

            card_positions = card.get("anyPositions") or []
            card_position = next(
                (
                    position
                    for position in card_positions
                    if position in POSITION_ORDER
                ),
                None,
            )

            result.append({
                "card_slug": card.get("slug") or "",
                "player_slug": player_slug,
                "player_name": player.get("displayName") or player_slug,
                "position": card_position,
                "rarity": "Limited",
                "season_year": CURRENT_SEASON_YEAR,
            })

        page_info = cards.get("pageInfo") or {}

        if not page_info.get("hasNextPage"):
            break

        after = page_info.get("endCursor")

        if not after:
            break

    return result


def match_owned_cards_to_league(
    owned_cards: List[dict],
    league_players: List[dict],
) -> Tuple[List[dict], List[dict]]:
    """
    Verknüpft die User-Karten mit den Spielern der ausgewählten Liga.

    Rückgabe:
    - jede einzelne passende Karte für das Excel-Blatt "Meine Karten"
    - eindeutige Spieler für die Best-Lineup-Berechnung
    """
    player_by_slug = {
        player["slug"]: player
        for player in league_players
    }

    matched_cards = []
    unique_players = {}

    for card in owned_cards:
        player = player_by_slug.get(card["player_slug"])

        if not player:
            continue

        actual_card_position = (
            card.get("position")
            if card.get("position") in POSITION_ORDER
            else player["position"]
        )

        row = {
            **card,
            "name": player["name"],
            "slug": player["slug"],
            "position": actual_card_position,
            "club_slug": player["club_slug"],
            "club_name": player["club_name"],
            "competition_slug": player.get("competition_slug"),
            "l40": float(player.get("l40", 0.0)),
            "starts_2026_27": int(player.get("starts_2026_27", 0)),
            "start_rate_2026_27": player.get("start_rate_2026_27"),
        }

        matched_cards.append(row)

        # Mehrere Karten desselben Spielers dürfen nicht dazu führen,
        # dass derselbe Spieler zweimal in einer Aufstellung verwendet wird.
        unique_players[player["slug"]] = {
            "slug": player["slug"],
            "name": player["name"],
            "position": actual_card_position,
            "club_slug": player["club_slug"],
            "club_name": player["club_name"],
            "competition_slug": player.get("competition_slug"),
            "l40": float(player.get("l40", 0.0)),
            "starts_2026_27": int(player.get("starts_2026_27", 0)),
        }

    matched_cards.sort(
        key=lambda item: (
            item["club_name"],
            POSITION_ORDER.index(item["position"]),
            item["name"],
        )
    )

    return matched_cards, list(unique_players.values())



def sorare_gameweek_window(date_raw: str) -> Tuple[str, str]:
    """
    Ordnet ein Spiel dem offiziellen Sorare-Football-Game-Week-Rhythmus zu:
    Dienstag 16:00 -> Freitag 16:00
    Freitag 16:00 -> Dienstag 16:00

    Für die Gruppierung wird Europe/Berlin verwendet. Dadurch werden die
    vier Contender-Ligen nicht nach ihren nationalen Spieltagsnummern,
    sondern nach demselben Sorare-Zeitfenster zusammengeführt.
    """
    from zoneinfo import ZoneInfo
    from datetime import timedelta

    tz = ZoneInfo("Europe/Berlin")
    dt = iso_to_datetime(date_raw).astimezone(tz)

    # Suche rückwärts nach der letzten Dienstag-/Freitag-Grenze um 16:00.
    candidates = []

    for days_back in range(0, 8):
        day = (dt - timedelta(days=days_back)).date()

        # weekday: Montag=0, Dienstag=1, Freitag=4
        if day.weekday() in (1, 4):
            boundary = datetime(
                day.year, day.month, day.day,
                16, 0, 0,
                tzinfo=tz,
            )

            if boundary <= dt:
                candidates.append(boundary)

    if not candidates:
        raise RuntimeError(
            f"Kein Sorare-Game-Week-Fenster für {date_raw} gefunden."
        )

    start_dt = max(candidates)

    if start_dt.weekday() == 1:  # Dienstag -> Freitag
        end_dt = start_dt + timedelta(days=3)
    else:  # Freitag -> Dienstag
        end_dt = start_dt + timedelta(days=4)

    key = start_dt.strftime("%Y-%m-%dT%H:%M:%S%z")
    label = (
        f"{start_dt.strftime('%d.%m. %H:%M')} – "
        f"{end_dt.strftime('%d.%m. %H:%M')}"
    )

    return key, label


def build_user_best_lineups(
    owned_players: List[dict],
    competition_slug: str,
) -> List[dict]:
    """
    Persönliche Rückschau.

    Bei Contender werden Spieler aus:
    - 2. Bundesliga
    - Ligue 2
    - Österreichischer Bundesliga
    - HNL

    GEMEINSAM in dasselbe Sorare-Game-Week-Fenster gelegt.

    Danach wird aus ALLEN eigenen passenden In-Season Limited 2026/27
    Spielern das bestmögliche Team gebildet:
    1 TW + 1 VER + 1 MID + 1 FW + 1 zusätzlicher Feldspieler.

    Kapitän = höchster L40 unter den tatsächlich gewerteten Feldspielern.
    """
    if not owned_players:
        return []

    players_by_club = defaultdict(list)

    for player in owned_players:
        players_by_club[
            (
                player["club_slug"],
                player["club_name"],
                player.get("competition_slug") or competition_slug,
            )
        ].append(player)

    gw_candidates = defaultdict(
        lambda: defaultdict(list)
    )
    gw_labels = {}

    total_clubs = len(players_by_club)

    print()
    print("=" * 70)
    print("PERSÖNLICHE BEST-LINEUPS")
    if competition_slug == CONTENDER_SLUG:
        print("CONTENDER = 2. BL + LIGUE 2 + ÖSTERREICH + HNL")
    print("Gruppierung nach Sorare Game Week (Di/Fr 16:00)")
    print("=" * 70)

    for index, (
        (club_slug, club_name, real_competition_slug),
        club_players,
    ) in enumerate(
        sorted(players_by_club.items(), key=lambda item: item[0][1]),
        start=1,
    ):
        try:
            games = get_club_matchday_scores_2026_27(
                club_slug,
                club_players,
                real_competition_slug,
            )
        except Exception as exc:
            print(
                f"[{index}/{total_clubs}] FEHLER persönliche Scores: "
                f"{club_name} -> {exc}"
            )
            continue

        for game in games:
            gw_key, gw_label = sorare_gameweek_window(game["date"])
            gw_labels[gw_key] = gw_label

            scores = game.get("scores") or {}

            for player in club_players:
                score = scores.get(player["slug"])

                if score is None:
                    continue

                gw_candidates[gw_key][
                    player["position"]
                ].append({
                    "slug": player["slug"],
                    "name": player["name"],
                    "club_name": player["club_name"],
                    "position": player["position"],
                    "score": float(score),
                    "l40": float(player.get("l40", 0.0)),
                    "competition_slug": real_competition_slug,
                })

        print(
            f"[{index:>2}/{total_clubs}] {club_name} | "
            f"{len(club_players)} eigene Spieler"
        )

    thresholds = LEAGUE_THRESHOLDS[competition_slug]
    results = []

    for number, gw_key in enumerate(sorted(gw_candidates.keys()), start=1):
        candidates_by_position = gw_candidates[gw_key]
        chosen = {}
        used = set()

        for position in POSITION_ORDER:
            candidates = candidates_by_position.get(position, [])

            if candidates:
                best = max(
                    candidates,
                    key=lambda item: item["score"],
                )
                chosen[position] = best
                used.add(best["slug"])
            else:
                chosen[position] = None

        extra_candidates = []

        for position in ("Defender", "Midfielder", "Forward"):
            for player in candidates_by_position.get(position, []):
                if player["slug"] not in used:
                    extra_candidates.append(player)

        extra = (
            max(extra_candidates, key=lambda item: item["score"])
            if extra_candidates
            else None
        )

        field_players = [
            chosen.get("Defender"),
            chosen.get("Midfielder"),
            chosen.get("Forward"),
            extra,
        ]
        field_players = [
            player for player in field_players
            if player is not None
        ]

        captain = (
            max(field_players, key=lambda item: item["l40"])
            if field_players
            else None
        )

        lineup = [
            chosen.get("Goalkeeper"),
            chosen.get("Defender"),
            chosen.get("Midfielder"),
            chosen.get("Forward"),
            extra,
        ]

        raw_sum = sum(
            player["score"]
            for player in lineup
            if player is not None
        )

        sum_a = 0.0

        for player in lineup:
            if player is None:
                continue

            if captain is not None and player["slug"] == captain["slug"]:
                sum_a += player["score"] * 1.58
            else:
                sum_a += player["score"] * 1.08

        mark_a = f"<{thresholds[0]}"

        for threshold in thresholds:
            if sum_a >= threshold:
                mark_a = str(threshold)

        results.append({
            "number": number,
            "date_text": gw_labels[gw_key],
            "gk": chosen.get("Goalkeeper"),
            "def": chosen.get("Defender"),
            "mid": chosen.get("Midfielder"),
            "fwd": chosen.get("Forward"),
            "extra": extra,
            "captain": captain,
            "raw_sum": raw_sum,
            "sum_a": sum_a,
            "mark_a": mark_a,
            "complete": all(player is not None for player in lineup),
        })

    return results


def avg(values: List[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def iso_to_datetime(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00"))



def club_has_current_season_limited_cards(club_slug: str) -> bool:
    """
    True nur dann, wenn Sorare für den Verein tatsächlich Limited-Karten
    der Saison 2026/27 führt.

    Dadurch verschwinden automatisch Vereine, die zwar sportlich in einer
    Competition spielen, aber in Sorare 27 keine aktuellen Karten haben.
    """
    query = """
    query ClubCurrentSeasonCards($clubSlug: String!) {
      cardsWhere(
        first: 1
        teamSlugs: [$clubSlug]
        rarities: [limited]
        seasonStartYears: [2026]
        sport: FOOTBALL
      ) {
        nodes {
          slug
          seasonYear
          rarityTyped
        }
      }
    }
    """

    data = graphql_request(query, {"clubSlug": club_slug})
    connection = data.get("cardsWhere") or {}
    return bool(connection.get("nodes"))


def filter_teams_with_current_season_cards(
    teams: Dict[str, str],
    competition_slug: str,
) -> Dict[str, str]:
    """
    Prüft jeden Verein der Competition gegen den echten Sorare-Kartenbestand
    für 2026/27 Limited.
    """
    if not teams:
        return {}

    print()
    print("=" * 70)
    print(f"2026/27 KARTEN-CHECK: {competition_slug}")
    print("=" * 70)

    valid = {}

    for index, (club_slug, club_name) in enumerate(
        sorted(teams.items(), key=lambda item: item[1]),
        start=1,
    ):
        try:
            has_cards = club_has_current_season_limited_cards(club_slug)
        except Exception as exc:
            # Bei einem API-Fehler lieber den Verein nicht heimlich entfernen.
            print(
                f"[{index:>2}/{len(teams)}] ? {club_name} | "
                f"Check fehlgeschlagen: {exc}"
            )
            valid[club_slug] = club_name
            continue

        if has_cards:
            valid[club_slug] = club_name
            print(
                f"[{index:>2}/{len(teams)}] ✅ {club_name} | "
                f"Limited 2026/27 vorhanden"
            )
        else:
            print(
                f"[{index:>2}/{len(teams)}] ❌ {club_name} | "
                f"keine Limited 2026/27 Karten -> entfernt"
            )

    print(
        f"✅ {len(valid)}/{len(teams)} Vereine mit aktuellen "
        f"Limited-Karten bleiben übrig."
    )

    return valid


def get_competition_teams(competition_slug: str) -> Dict[str, str]:
    query = """
    query CompetitionTeams($slug: String!, $after: String) {
      football {
        competition(slug: $slug) {
          teams(first: 50, after: $after) {
            pageInfo { hasNextPage endCursor }
            nodes { slug name }
          }
        }
      }
    }
    """
    teams = {}
    after = None
    while True:
        data = graphql_request(query, {"slug": competition_slug, "after": after})
        competition = data["football"].get("competition")
        if not competition:
            raise RuntimeError(f"Competition '{competition_slug}' wurde nicht gefunden.")
        connection = competition["teams"]
        for team in connection.get("nodes") or []:
            if team and team.get("slug"):
                teams[team["slug"]] = team.get("name") or team["slug"]
        page_info = connection.get("pageInfo") or {}
        if not page_info.get("hasNextPage"):
            break
        after = page_info.get("endCursor")
        if not after:
            break
    return teams


def get_league_players(
    competition_slug: str,
    competition_teams: Dict[str, str],
) -> List[dict]:
    query = """
    query CompetitionPlayers($slug: String!, $after: String) {
      football {
        competition(slug: $slug) {
          orderedPlayers(first: 50, after: $after) {
            pageInfo { hasNextPage endCursor }
            nodes {
              slug
              displayName
              position
              activeClub { slug name }
            }
          }
        }
      }
    }
    """
    # Name -> aktueller Competition-Team-Slug
    team_slug_by_name = {
        team_name.casefold(): team_slug
        for team_slug, team_name in competition_teams.items()
    }
    allowed_team_slugs = set(competition_teams.keys())

    players = []
    filtered_out = 0
    corrected_transfers = []
    after = None

    while True:
        data = graphql_request(query, {"slug": competition_slug, "after": after})
        competition = data["football"].get("competition")
        if not competition:
            raise RuntimeError(f"Competition '{competition_slug}' wurde nicht gefunden.")

        connection = competition["orderedPlayers"]

        for player in connection.get("nodes") or []:
            if not player:
                continue
            club = player.get("activeClub")
            position = player.get("position")

            # Die Player-Position auf der Profilseite ist NICHT unsere
            # endgültige Position. Entscheidend ist später die Position auf
            # der 2026/27 Sorare-Karte.
            player_name = player.get("displayName") or player["slug"]

            club_slug = (club or {}).get("slug") or ""
            club_name = (club or {}).get("name") or "Unbekannter Verein"

            # Frische Transfers korrigieren, falls Sorare noch den alten
            # activeClub zurückgibt.
            override_team_name = PLAYER_CLUB_OVERRIDES.get(player_name)

            if override_team_name:
                override_slug = team_slug_by_name.get(
                    override_team_name.casefold()
                )

                if override_slug:
                    old_club_name = club_name
                    club_slug = override_slug
                    club_name = competition_teams[override_slug]

                    corrected_transfers.append(
                        f"{player_name}: {old_club_name} -> {club_name}"
                    )

            # Nach möglicher Transfer-Korrektur nur aktuelle Ligavereine zulassen.
            if club_slug not in allowed_team_slugs:
                filtered_out += 1
                continue

            players.append({
                "slug": player["slug"],
                "name": player_name,
                "position": position,
                "club_slug": club_slug,
                "club_name": club_name,
            })

        page_info = connection.get("pageInfo") or {}
        if not page_info.get("hasNextPage"):
            break
        after = page_info.get("endCursor")
        if not after:
            break

    print(f"✅ {len(players)} gültige Spieler geladen.")
    if filtered_out:
        print(f"✅ {filtered_out} Spieler außerhalb der Liga herausgefiltert.")

    if corrected_transfers:
        print()
        print("🔄 Manuell korrigierte aktuelle Vereine:")
        for correction in sorted(set(corrected_transfers)):
            print(f"   - {correction}")

    return players



def get_current_card_position_and_l40(
    player_slug: str,
) -> Tuple[Optional[str], float]:
    """
    Holt für einen Spieler in EINER Anfrage:
    - eine echte Limited-Karte aus 2026/27
    - deren dauerhafte Kartenposition
    - L40 genau für DIESE Kartenposition

    Wichtig: Sorare-Spielerprofilposition und Kartenposition können
    voneinander abweichen. Für unseren Analyzer zählt die Kartenposition.
    """
    query = """
    query PlayerCardPositionAndScores($slug: String!) {
      anyPlayer(slug: $slug) {
        ... on Player {
          currentCards: anyCards(
            first: 1
            rarities: [limited]
            seasonStartYears: [2026]
            sport: FOOTBALL
          ) {
            nodes {
              slug
              seasonYear
              rarityTyped
              anyPositions
              anyTeam {
                slug
                name
              }
            }
          }

          gkScores: rawPlayerGameScores(
            last: 40
            position: Goalkeeper
          )
          defScores: rawPlayerGameScores(
            last: 40
            position: Defender
          )
          midScores: rawPlayerGameScores(
            last: 40
            position: Midfielder
          )
          fwdScores: rawPlayerGameScores(
            last: 40
            position: Forward
          )
        }
      }
    }
    """

    data = graphql_request(query, {"slug": player_slug})
    player = data.get("anyPlayer")

    if not player:
        return None, 0.0

    connection = player.get("currentCards") or {}
    nodes = connection.get("nodes") or []

    if not nodes:
        # Keine Limited-Karte dieser Saison -> nicht für unseren
        # In-Season-2026/27-Analyzer verwenden.
        return None, 0.0

    card = nodes[0] or {}
    positions = card.get("anyPositions") or []

    card_position = next(
        (
            position
            for position in positions
            if position in POSITION_ORDER
        ),
        None,
    )

    if card_position is None:
        return None, 0.0

    score_field_by_position = {
        "Goalkeeper": "gkScores",
        "Defender": "defScores",
        "Midfielder": "midScores",
        "Forward": "fwdScores",
    }

    score_field = score_field_by_position[card_position]

    scores = [
        float(score)
        for score in (player.get(score_field) or [])
        if score is not None
    ]

    return card_position, avg(scores)


def add_l40(players: List[dict]):
    """
    Prüft JEDEN Spieler gegen seine echte 2026/27-Limited-Karte.

    - Position wird von der Karte übernommen.
    - Spieler ohne aktuelle Limited-Karte werden entfernt.
    - L40 wird für genau diese Kartenposition berechnet.
    """
    print("\n" + "=" * 70)
    print("2026/27 KARTENPOSITIONEN + L40")
    print("=" * 70 + "\n")

    original_total = len(players)
    valid_players = []
    changed_positions = []

    for index, player in enumerate(players, start=1):
        old_position = player.get("position")

        try:
            card_position, l40 = get_current_card_position_and_l40(
                player["slug"]
            )
        except Exception as exc:
            print(
                f"[{index}/{original_total}] FEHLER KARTENCHECK: "
                f"{player['name']} -> {exc}"
            )
            # Bei einem echten API-Fehler entfernen wir den Spieler nicht
            # stillschweigend. Seine bisherige Position bleibt als Fallback.
            if old_position in POSITION_ORDER:
                player["l40"] = 0.0
                valid_players.append(player)
            continue

        if card_position is None:
            print(
                f"[{index:>3}/{original_total}] ❌ "
                f"{player['club_name']} | {player['name']} | "
                f"keine Limited 2026/27 Karte -> entfernt"
            )
            continue

        player["position"] = card_position
        player["l40"] = l40
        valid_players.append(player)

        if old_position != card_position:
            changed_positions.append(
                (
                    player["name"],
                    old_position,
                    card_position,
                )
            )

        old_label = POSITION_LABELS.get(
            old_position,
            str(old_position or "-"),
        )
        new_label = POSITION_LABELS[card_position]

        change_text = (
            f" | Profil {old_label} -> Karte {new_label}"
            if old_position != card_position
            else ""
        )

        print(
            f"[{index:>3}/{original_total}] "
            f"{player['club_name']} | "
            f"{new_label} | "
            f"{player['name']} | "
            f"L40 {player['l40']:.2f}"
            f"{change_text}"
        )

    # Liste in-place ersetzen, damit alle späteren Schritte nur noch
    # tatsächlich verfügbare aktuelle Karten verwenden.
    players[:] = valid_players

    print()
    print(
        f"✅ {len(players)}/{original_total} Spieler mit "
        f"Limited 2026/27 Karte verbleiben."
    )

    if changed_positions:
        print()
        print("🔄 Abweichende Profil-/Kartenpositionen korrigiert:")
        for name, old_position, new_position in changed_positions:
            print(
                f"   - {name}: "
                f"{POSITION_LABELS.get(old_position, old_position)} -> "
                f"{POSITION_LABELS[new_position]}"
            )


def get_analysis_competition_slugs(competition_slug: str) -> List[str]:
    if competition_slug == CONTENDER_SLUG:
        return list(CONTENDER_COMPETITIONS)
    return [competition_slug]


def get_combined_competition_teams(
    competition_slug: str,
) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Lädt Teams für eine normale Liga oder für alle vier Contender-Ligen.

    Rückgabe:
    - alle Team-Slugs -> Teamname
    - Team-Slug -> echte Sorare-Competition
    """
    all_teams = {}
    team_competition = {}

    for real_slug in get_analysis_competition_slugs(competition_slug):
        teams = get_competition_teams(real_slug)
        teams = filter_teams_with_current_season_cards(
            teams,
            real_slug,
        )

        # Ligue 2 ist ein Sonderfall:
        # Nur diese fünf Sorare-Collections 2026/27 gehören in unseren
        # Contender-Analyzer. Alle anderen sportlichen Ligue-2-Vereine
        # werden ausgeschlossen.
        if real_slug == "ligue-2-fr":
            before_count = len(teams)
            teams = {
                team_slug: team_name
                for team_slug, team_name in teams.items()
                if team_name in LIGUE_2_SORARE_2026_27_CLUBS
            }

            print()
            print("=" * 70)
            print("LIGUE 2 – FESTER SORARE-27-KARTENFILTER")
            print("=" * 70)

            for team_name in sorted(teams.values()):
                print(f"✅ {team_name}")

            print(
                f"✅ Ligue 2: {len(teams)} Vereine im Analyzer "
                f"(vor Filter: {before_count})"
            )

            if len(teams) != 5:
                print(
                    "⚠️ WARNUNG: Es wurden nicht alle 5 erwarteten "
                    "Ligue-2-Sorare-Clubs gefunden. Bitte die oben "
                    "angezeigten Vereinsnamen prüfen."
                )

        for team_slug, team_name in teams.items():
            all_teams[team_slug] = team_name
            team_competition[team_slug] = real_slug

    return all_teams, team_competition


def get_combined_league_players(
    competition_slug: str,
    competition_teams: Dict[str, str],
    team_competition: Dict[str, str],
) -> List[dict]:
    """
    Normale Liga: unverändert.
    Contender: Spieler aus 2. Bundesliga, Ligue 2, Österreich und HNL
    werden zu EINEM Spielerpool zusammengeführt.
    """
    if competition_slug != CONTENDER_SLUG:
        players = get_league_players(
            competition_slug,
            competition_teams,
        )
        for player in players:
            player["competition_slug"] = competition_slug
        return players

    result = []
    seen = set()

    for real_slug in CONTENDER_COMPETITIONS:
        real_teams = {
            team_slug: team_name
            for team_slug, team_name in competition_teams.items()
            if team_competition.get(team_slug) == real_slug
        }

        players = get_league_players(real_slug, real_teams)

        for player in players:
            if player["slug"] in seen:
                continue

            seen.add(player["slug"])
            player["competition_slug"] = real_slug
            result.append(player)

    return result


def add_combined_starter_stats(
    players: List[dict],
    competition_slug: str,
):
    """
    Startelfquote bleibt liga-echt:
    Ein Spieler aus Österreich wird nur mit österreichischen Ligaspielen
    berechnet, ein Spieler aus der 2. Bundesliga nur mit deren Spielen usw.
    """
    if competition_slug != CONTENDER_SLUG:
        add_starter_stats(players, competition_slug)
        return

    for real_slug in CONTENDER_COMPETITIONS:
        subset = [
            player
            for player in players
            if player.get("competition_slug") == real_slug
        ]

        if subset:
            add_starter_stats(subset, real_slug)


def get_club_starter_stats_2026_27(
    club_slug: str,
    competition_slug: str,
) -> Tuple[Dict[str, int], int]:
    """
    Zählt Startelf-Einsätze NUR in der ausgewählten Liga.

    Standard für die europäischen Ligen:
    - alle Ligaspiele der Saison 2026/27 innerhalb unseres Saisonfensters

    Sonderregel MLS:
    - alle absolvierten MLS-Ligaspiele der Saison 2026
      mit verfügbaren Startelfdaten
    - die komplette bisherige MLS-Saison seit Saisonstart wird für
      die 70/30-Spielerauswahl berücksichtigt

    Keine Pokal-, Champions-League-, Europa-League- oder
    Freundschaftsspiele.
    """
    query = """
    query CompetitionPastGames(
      $competitionSlug: String!
      $after: String
    ) {
      football {
        competition(slug: $competitionSlug) {
          pastGames(first: 50, after: $after) {
            pageInfo {
              hasNextPage
              endCursor
            }
            nodes {
              date
              homeTeam {
                slug
              }
              awayTeam {
                slug
              }
              homeFormation {
                startingLineupAvailable
                startingLineup {
                  slug
                }
              }
              awayFormation {
                startingLineupAvailable
                startingLineup {
                  slug
                }
              }
            }
          }
        }
      }
    }
    """

    after = None
    season_start = iso_to_datetime(SEASON_START)
    season_end = iso_to_datetime(SEASON_END)
    now = datetime.now(timezone.utc)

    # Wir sammeln alle passenden Ligaspiele des Vereins.
    # Für MLS zählt die komplette bisherige Saison 2026.
    eligible_club_games = []

    while True:
        data = graphql_request(
            query,
            {
                "competitionSlug": competition_slug,
                "after": after,
            },
        )

        competition = data["football"].get("competition")

        if not competition:
            break

        connection = competition["pastGames"]
        page_dates = []

        for game in connection.get("nodes") or []:
            game_date_raw = game.get("date")

            if not game_date_raw:
                continue

            game_date = iso_to_datetime(game_date_raw)
            page_dates.append(game_date)

            # MLS läuft im Kalenderjahr 2026.
            # Für MLS nehmen wir alle bisherigen Spiele aus 2026 als Pool,
            # bevor wir daraus die letzten 10 auswählen.
            if competition_slug == "mlspa":
                if game_date.year != CURRENT_SEASON_YEAR:
                    continue
                if game_date > now:
                    continue
            else:
                if game_date < season_start:
                    continue
                if game_date > season_end or game_date > now:
                    continue

            home_team = game.get("homeTeam") or {}
            away_team = game.get("awayTeam") or {}

            if home_team.get("slug") == club_slug:
                formation = game.get("homeFormation") or {}
            elif away_team.get("slug") == club_slug:
                formation = game.get("awayFormation") or {}
            else:
                continue

            if not formation.get("startingLineupAvailable"):
                continue

            player_slugs = []

            for group in formation.get("startingLineup") or []:
                if isinstance(group, dict):
                    group = [group]

                for player in group or []:
                    if not player:
                        continue

                    player_slug = player.get("slug")

                    if player_slug:
                        player_slugs.append(player_slug)

            eligible_club_games.append(
                {
                    "date": game_date,
                    "player_slugs": player_slugs,
                }
            )

        page_info = connection.get("pageInfo") or {}

        if not page_info.get("hasNextPage"):
            break

        # Bei europäischen Ligen können wir aufhören, sobald die komplette
        # Seite vor unserem Saisonstart liegt. Für MLS brauchen wir dagegen
        # den kompletten 2026-Pool, damit die letzten 10 sicher stimmen.
        if (
            competition_slug != "mlspa"
            and page_dates
            and max(page_dates) < season_start
        ):
            break

        after = page_info.get("endCursor")

        if not after:
            break

    eligible_club_games.sort(
        key=lambda item: item["date"],
        reverse=True,
    )

    counts = defaultdict(int)

    for game in eligible_club_games:
        for player_slug in game["player_slugs"]:
            counts[player_slug] += 1

    eligible_games = len(eligible_club_games)

    return dict(counts), eligible_games



def add_starter_stats(players: List[dict], competition_slug: str):
    players_by_club = defaultdict(list)
    for player in players:
        players_by_club[(player["club_slug"], player["club_name"])].append(player)

    print("\n" + "=" * 70)
    if competition_slug == "mlspa":
        print("STARTELF-EINSÄTZE + STARTELFQUOTE | MLS: GESAMTE SAISON 2026")
    else:
        print("STARTELF-EINSÄTZE + STARTELFQUOTE 2026/27")
    print("=" * 70 + "\n")

    total = len(players_by_club)

    for index, ((club_slug, club_name), club_players) in enumerate(
        sorted(players_by_club.items(), key=lambda item: item[0][1]),
        start=1,
    ):
        try:
            counts, eligible_games = get_club_starter_stats_2026_27(club_slug, competition_slug)
        except Exception as exc:
            print(f"[{index}/{total}] FEHLER STARTELF: {club_name} -> {exc}")
            counts, eligible_games = {}, 0

        for player in club_players:
            starts = counts.get(player["slug"], 0)
            player["starts_2026_27"] = starts
            player["eligible_games_2026_27"] = eligible_games
            player["start_rate_2026_27"] = (
                (starts / eligible_games) * 100.0
                if eligible_games > 0
                else None
            )

        print(
            f"[{index:>2}/{total}] {club_name} | "
            f"{eligible_games} Spiele mit Startelfdaten"
        )


def calculate_selection_score(player: dict) -> float:
    l40 = float(player.get("l40", 0.0))
    start_rate = player.get("start_rate_2026_27")
    if start_rate is None:
        return l40
    return l40 * L40_WEIGHT + float(start_rate) * START_RATE_WEIGHT


def build_selection(players: List[dict]):
    grouped = defaultdict(lambda: defaultdict(list))

    for player in players:
        player["selection_score"] = calculate_selection_score(player)
        grouped[player["club_name"]][player["position"]].append(player)

    selected = defaultdict(lambda: defaultdict(list))

    print("\n" + "=" * 90)
    print("TOP-8 AUSWAHL: 70 % L40 + 30 % STARTELFQUOTE")
    print("=" * 90)

    for club_name in sorted(grouped.keys()):
        print(f"\n{club_name}\n" + "-" * 90)

        for position in POSITION_ORDER:
            candidates = grouped[club_name].get(position, [])
            candidates.sort(
                key=lambda item: (
                    item.get("selection_score", 0.0),
                    item.get("l40", 0.0),
                    item.get("starts_2026_27", 0),
                ),
                reverse=True,
            )

            chosen = candidates[:2]
            selected[club_name][position] = chosen

            for priority, player in enumerate(chosen, start=1):
                start_rate = player.get("start_rate_2026_27")
                rate_text = "-" if start_rate is None else f"{start_rate:.0f}%"
                print(
                    f"{POSITION_LABELS[position]:<4} | "
                    f"Prio {priority} | "
                    f"{player['name']:<27} | "
                    f"L40 {player['l40']:>5.1f} | "
                    f"Starts {player['starts_2026_27']:>2}/"
                    f"{player['eligible_games_2026_27']:<2} | "
                    f"Quote {rate_text:>4} | "
                    f"Auswahlwert {player['selection_score']:>5.1f}"
                )

    return selected


def get_cheapest_inseason_limited_price(player_slug: str) -> Optional[float]:
    query = """
    query PlayerOffers($playerSlug: String!, $after: String) {
      tokens {
        liveSingleSaleOffers(
          first: 50
          after: $after
          playerSlug: $playerSlug
          sport: FOOTBALL
        ) {
          pageInfo { hasNextPage endCursor }
          nodes {
            senderSide {
              anyCards { rarityTyped seasonYear }
            }
            receiverSide {
              amounts { eurCents }
            }
          }
        }
      }
    }
    """
    after = None
    prices = []

    while True:
        data = graphql_request(
            query,
            {"playerSlug": player_slug, "after": after},
        )
        connection = data["tokens"]["liveSingleSaleOffers"]

        for offer in connection.get("nodes") or []:
            cards = (offer.get("senderSide") or {}).get("anyCards") or []
            valid_card = any(
                str(card.get("rarityTyped") or "").lower() == "limited"
                and card.get("seasonYear") == CURRENT_SEASON_YEAR
                for card in cards
            )
            if not valid_card:
                continue

            eur_cents = (
                (offer.get("receiverSide") or {})
                .get("amounts", {})
                .get("eurCents")
            )
            if eur_cents is None:
                continue

            try:
                prices.append(float(eur_cents) / 100.0)
            except (TypeError, ValueError):
                pass

        page_info = connection.get("pageInfo") or {}
        if not page_info.get("hasNextPage"):
            break
        after = page_info.get("endCursor")
        if not after:
            break

    return min(prices) if prices else None


def add_prices(selected):
    all_selected = []
    for club_name in sorted(selected.keys()):
        for position in POSITION_ORDER:
            all_selected.extend(selected[club_name].get(position, []))

    print("\n" + "=" * 70)
    print("AKTUELLE IN-SEASON LIMITED 2026/27 PREISE")
    print("=" * 70 + "\n")

    total = len(all_selected)

    for index, player in enumerate(all_selected, start=1):
        try:
            player["price"] = get_cheapest_inseason_limited_price(player["slug"])
        except Exception as exc:
            print(f"[{index}/{total}] FEHLER PREIS: {player['name']} -> {exc}")
            player["price"] = None

        price_text = (
            f"{player['price']:.2f} €"
            if player["price"] is not None
            else "kein Angebot"
        )
        print(
            f"[{index:>3}/{total}] {player['club_name']} | "
            f"{POSITION_LABELS[player['position']]} | "
            f"{player['name']} | {price_text}"
        )



# ============================================================
# Spieltage + Sorare Scores 2026/27
# ============================================================

def get_club_matchday_scores_2026_27(
    club_slug: str,
    selected_players: List[dict],
    competition_slug: str,
) -> List[dict]:
    """
    Lädt die vergangenen Spiele DIREKT aus der ausgewählten Competition.

    Dadurch kommen wirklich nur Ligaspiele hinein:
    - Bundesliga -> nur Bundesliga
    - Premier League -> nur Premier League
    usw.

    Internationale Spiele und Pokalspiele sind damit automatisch draußen.
    """
    if not selected_players:
        return []

    variable_defs = []
    score_fields = []
    variables = {
        "competitionSlug": competition_slug,
        "after": None,
    }

    alias_to_player = {}

    for index, player in enumerate(selected_players):
        var_name = f"p{index}"
        alias = f"score{index}"

        variable_defs.append(f"${var_name}: String!")
        card_position = player["position"]

        score_fields.append(
            f"""
            {alias}: playerGameScore(
              playerSlug: ${var_name}
              position: {card_position}
            ) {{
              score
            }}
            """
        )

        variables[var_name] = player["slug"]
        alias_to_player[alias] = player

    extra_defs = ", " + ", ".join(variable_defs) if variable_defs else ""
    score_fields_text = "\n".join(score_fields)

    query = f"""
    query CompetitionPastGames(
      $competitionSlug: String!
      $after: String
      {extra_defs}
    ) {{
      football {{
        competition(slug: $competitionSlug) {{
          pastGames(first: 50, after: $after) {{
            pageInfo {{
              hasNextPage
              endCursor
            }}
            nodes {{
              id
              date
              homeTeam {{
                slug
                name
              }}
              awayTeam {{
                slug
                name
              }}
              {score_fields_text}
            }}
          }}
        }}
      }}
    }}
    """

    after = None
    results = []

    season_start = iso_to_datetime(SEASON_START)
    season_end = iso_to_datetime(SEASON_END)
    now = datetime.now(timezone.utc)

    while True:
        variables["after"] = after
        data = graphql_request(query, variables)

        competition = data["football"].get("competition")

        if not competition:
            break

        connection = competition["pastGames"]
        page_dates = []

        for game in connection.get("nodes") or []:
            date_raw = game.get("date")

            if not date_raw:
                continue

            game_date = iso_to_datetime(date_raw)
            page_dates.append(game_date)

            # MLS: komplette bisherige Saison 2026.
            # Europäische Ligen: Saisonfenster 2026/27.
            if competition_slug == "mlspa":
                if game_date.year != CURRENT_SEASON_YEAR:
                    continue
                if game_date > now:
                    continue
            else:
                if game_date < season_start:
                    continue
                if game_date > season_end or game_date > now:
                    continue

            home = game.get("homeTeam") or {}
            away = game.get("awayTeam") or {}

            # Nur Spiele des Vereins, den wir gerade auswerten.
            if home.get("slug") == club_slug:
                opponent = away.get("name") or away.get("slug") or "-"
                ha = "H"
            elif away.get("slug") == club_slug:
                opponent = home.get("name") or home.get("slug") or "-"
                ha = "A"
            else:
                continue

            player_scores = {}

            for alias, player in alias_to_player.items():
                score_node = game.get(alias)

                if not score_node:
                    continue

                score = score_node.get("score")

                if score is None:
                    continue

                try:
                    player_scores[player["slug"]] = float(score)
                except (TypeError, ValueError):
                    pass

            if not player_scores:
                continue

            results.append({
                "game_id": str(game.get("id") or ""),
                "date": date_raw,
                "opponent": opponent,
                "ha": ha,
                "scores": player_scores,
            })

        page_info = connection.get("pageInfo") or {}

        if not page_info.get("hasNextPage"):
            break

        # Europäische Ligen dürfen vor Saisonstart abbrechen.
        # MLS muss weiter zurückblättern, damit alle Spiele aus 2026
        # berücksichtigt werden.
        if (
            competition_slug != "mlspa"
            and page_dates
            and max(page_dates) < season_start
        ):
            break

        after = page_info.get("endCursor")

        if not after:
            break

    results.sort(key=lambda game: iso_to_datetime(game["date"]))
    return results



def build_matchday_rows(selected, competition_slug: str) -> List[dict]:
    """
    Pro Verein/Spiel:
    - bester Score der 2 TW
    - bester Score der 2 VER
    - bester Score der 2 MID
    - bester Score der 2 FW
    - bester noch nicht verwendeter Feldspieler aus VER/MID/FW

    Kapitän:
    - Torwart ausgeschlossen
    - von den 4 gewerteten Feldspielern wird der mit dem höchsten L40
      zum Kapitän
    - Kapitän: Score x 1,58
    - alle anderen 4: Score x 1,08
    """
    rows = []

    print()
    print("=" * 70)
    if competition_slug == "mlspa":
        print("SPIELTAGE + SORARE SCORES | MLS GESAMTE SAISON 2026")
    else:
        print("SPIELTAGE + SORARE SCORES 2026/27")
    print("Kapitän = höchster L40 der gewerteten Feldspieler")
    print("=" * 70)
    print()

    total_clubs = len(selected)

    for club_index, club_name in enumerate(sorted(selected.keys()), start=1):
        club_slug = None
        selected_players = []

        for position in POSITION_ORDER:
            for player in selected[club_name].get(position, []):
                club_slug = club_slug or player["club_slug"]
                selected_players.append(player)

        if not club_slug or not selected_players:
            continue

        real_competition_slug = (
            selected_players[0].get("competition_slug")
            or competition_slug
        )

        try:
            club_games = get_club_matchday_scores_2026_27(
                club_slug,
                selected_players,
                real_competition_slug,
            )
        except Exception as exc:
            print(
                f"[{club_index}/{total_clubs}] FEHLER Spieltage: "
                f"{club_name} -> {exc}"
            )
            continue

        for number, game in enumerate(club_games, start=1):
            scores = game["scores"]
            used = set()
            chosen_by_position = {}

            # Je Position bester tatsächlich erzielter Score.
            for position in POSITION_ORDER:
                candidates = []

                for player in selected[club_name].get(position, []):
                    slug = player["slug"]

                    if slug not in scores:
                        continue

                    candidates.append({
                        "slug": slug,
                        "score": scores[slug],
                        "l40": player.get("l40", 0.0),
                        "name": player["name"],
                    })

                if candidates:
                    chosen = max(
                        candidates,
                        key=lambda item: item["score"],
                    )
                    chosen_by_position[position] = chosen
                    used.add(chosen["slug"])
                else:
                    chosen_by_position[position] = None

            # Bester übriger Feldspieler.
            extra_candidates = []

            for position in ("Defender", "Midfielder", "Forward"):
                for player in selected[club_name].get(position, []):
                    slug = player["slug"]

                    if slug in used:
                        continue

                    if slug not in scores:
                        continue

                    extra_candidates.append({
                        "slug": slug,
                        "score": scores[slug],
                        "l40": player.get("l40", 0.0),
                        "name": player["name"],
                    })

            extra_player = (
                max(
                    extra_candidates,
                    key=lambda item: item["score"],
                )
                if extra_candidates
                else None
            )

            valued_players = []

            for position in POSITION_ORDER:
                chosen = chosen_by_position[position]

                if chosen is not None:
                    valued_players.append({
                        **chosen,
                        "position": position,
                    })
                else:
                    valued_players.append({
                        "slug": "",
                        "score": 0.0,
                        "l40": 0.0,
                        "name": "-",
                        "position": position,
                    })

            if extra_player is not None:
                valued_players.append({
                    **extra_player,
                    "position": "ExtraField",
                })
            else:
                valued_players.append({
                    "slug": "",
                    "score": 0.0,
                    "l40": 0.0,
                    "name": "-",
                    "position": "ExtraField",
                })

            # Kapitän ausschließlich aus echten gewerteten Feldspielern.
            captain_candidates = [
                player
                for player in valued_players
                if player["position"] != "Goalkeeper"
                and player["slug"]
            ]

            if captain_candidates:
                captain = max(
                    captain_candidates,
                    key=lambda item: item["l40"],
                )
            else:
                captain = None

            raw_sum = sum(
                player["score"]
                for player in valued_players
            )

            sum_a = 0.0

            for player in valued_players:
                if captain is not None and player is captain:
                    sum_a += player["score"] * 1.58
                else:
                    sum_a += player["score"] * 1.08

            thresholds = LEAGUE_THRESHOLDS[competition_slug]
            reached = [
                threshold
                for threshold in thresholds
                if sum_a >= threshold
            ]

            mark_a = max(reached) if reached else f"<{thresholds[0]}"

            gk = chosen_by_position["Goalkeeper"]
            de = chosen_by_position["Defender"]
            mi = chosen_by_position["Midfielder"]
            fw = chosen_by_position["Forward"]

            rows.append({
                "club_name": club_name,
                "number": number,
                "date": iso_to_datetime(game["date"]).date(),
                "opponent": game["opponent"],
                "ha": game["ha"],

                "gk_name": gk["name"] if gk else "-",
                "gk": gk["score"] if gk else 0.0,

                "def_name": de["name"] if de else "-",
                "def": de["score"] if de else 0.0,

                "mid_name": mi["name"] if mi else "-",
                "mid": mi["score"] if mi else 0.0,

                "fwd_name": fw["name"] if fw else "-",
                "fwd": fw["score"] if fw else 0.0,

                "field_name": extra_player["name"] if extra_player else "-",
                "field": extra_player["score"] if extra_player else 0.0,

                "raw_sum": raw_sum,
                "sum_a": sum_a,
                "mark_a": mark_a,
                "captain_name": captain["name"] if captain else "-",
                "captain_l40": captain["l40"] if captain else 0.0,
            })

        print(
            f"[{club_index:>2}/{total_clubs}] "
            f"{club_name}: {len(club_games)} Spiele"
        )

    return rows



def style_header(ws, row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.font = Font(bold=True, color="1E1E1E")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=THIN_BORDER, right=THIN_BORDER,
            top=THIN_BORDER, bottom=THIN_BORDER,
        )


def style_body_cell(cell, centered=False, bold=False):
    cell.fill = PatternFill("solid", fgColor=COLOR_DARK)
    cell.font = Font(color=COLOR_TEXT, bold=bold)
    cell.alignment = Alignment(
        horizontal="center" if centered else "left",
        vertical="center",
    )
    cell.border = Border(
        left=THIN_BORDER, right=THIN_BORDER,
        top=THIN_BORDER, bottom=THIN_BORDER,
    )


def create_excel(
    selected,
    league_name: str,
    competition_slug: str,
    matchday_rows: List[dict],
    sorare_slug: str,
    owned_cards: List[dict],
    user_best_lineups: List[dict],
) -> str:
    wb = Workbook()
    ws_players = wb.active
    ws_players.title = "Spielerauswahl"
    ws_clubs = wb.create_sheet("Vereinsübersicht")
    ws_games = wb.create_sheet("Spieltage")
    ws_details = wb.create_sheet("Spieltagsdetails")
    ws_my_cards = wb.create_sheet("Meine Karten")
    ws_my_lineup = wb.create_sheet("Mein bestes Lineup")

    for ws in (
        ws_players,
        ws_clubs,
        ws_games,
        ws_details,
        ws_my_cards,
        ws_my_lineup,
    ):
        ws.sheet_view.showGridLines = False

    player_headers = [
        "Verein",
        "Position",
        "Priorität",
        "Spieler",
        "L40",
        "Startelf-Einsätze 26/27",
        "Aktueller In-Season Preis 2026/27 (€)",
    ]

    ws_players.append(player_headers)
    style_header(ws_players, 1, 1, 7)
    ws_players.row_dimensions[1].height = 34
    ws_players.freeze_panes = "A2"

    row = 2

    for club_name in sorted(selected.keys()):
        first_row = row
        selected_count = 0
        price_count = 0

        for position in POSITION_ORDER:
            for priority, player in enumerate(
                selected[club_name].get(position, []), start=1
            ):
                selected_count += 1
                if player.get("price") is not None:
                    price_count += 1

                values = [
                    club_name,
                    POSITION_LABELS[position],
                    priority,
                    player["name"],
                    round(player.get("l40", 0.0), 1),
                    player.get("starts_2026_27", 0),
                    player.get("price"),
                ]

                for col, value in enumerate(values, start=1):
                    ws_players.cell(row, col, value)
                    style_body_cell(
                        ws_players.cell(row, col),
                        centered=(col in (2, 3, 5, 6, 7)),
                        bold=(priority == 1 and col in (1, 2, 3, 4)),
                    )

                ws_players.cell(row, 7).number_format = '#,##0.00 [$€-de-DE]'
                row += 1

        if selected_count == 8 and price_count == 8:
            label = "Kosten aller 8 Karten"
        elif selected_count == 8:
            label = f"Kosten: nur {price_count}/8 Preise verfügbar"
        else:
            label = f"Achtung: nur {selected_count}/8 Spieler gefunden"

        ws_players.cell(row, 1, label)
        ws_players.cell(row, 7, f"=SUM(G{first_row}:G{row - 1})")

        for col in range(1, 8):
            cell = ws_players.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL)
            cell.font = Font(color=COLOR_WHITE, bold=True)
            cell.border = Border(top=GOLD_BORDER, bottom=GOLD_BORDER)
            cell.alignment = Alignment(
                horizontal="center" if col in (3, 5, 6, 7) else "left",
                vertical="center",
            )

        ws_players.cell(row, 7).number_format = '#,##0.00 [$€-de-DE]'
        row += 1

    for col, width in {
        "A": 31, "B": 12, "C": 11, "D": 27,
        "E": 10, "F": 23, "G": 32,
    }.items():
        ws_players.column_dimensions[col].width = width

    ws_players.auto_filter.ref = f"A1:G{max(1, row - 1)}"

    thresholds = LEAGUE_THRESHOLDS[competition_slug]

    club_headers = [
        "Verein",
        "Kosten 8 Karten (€)",
        "Spieltage",
        "Ø A",
        "Bester Tag A",
    ] + [f"A ≥{threshold}" for threshold in thresholds]

    ws_clubs.append(club_headers)
    style_header(ws_clubs, 1, 1, len(club_headers))
    ws_clubs.freeze_panes = "A2"

    club_row = 2

    matchdays_by_club = defaultdict(list)
    for item in matchday_rows:
        matchdays_by_club[item["club_name"]].append(item)

    club_summary_rows = []

    for club_name in sorted(selected.keys()):
        prices = [
            player["price"]
            for position in POSITION_ORDER
            for player in selected[club_name].get(position, [])
            if player.get("price") is not None
        ]

        club_matchdays = matchdays_by_club.get(club_name, [])
        sum_a_values = [
            float(item["sum_a"])
            for item in club_matchdays
            if item.get("sum_a") is not None
        ]

        games_count = len(sum_a_values)
        average_a = (
            sum(sum_a_values) / games_count
            if games_count else None
        )
        best_a = max(sum_a_values) if sum_a_values else None

        threshold_counts = {
            threshold: sum(
                1 for value in sum_a_values
                if value >= threshold
            )
            for threshold in thresholds
        }

        club_summary_rows.append({
            "club_name": club_name,
            "cost": sum(prices) if prices else None,
            "games_count": games_count,
            "average_a": average_a,
            "best_a": best_a,
            "threshold_counts": threshold_counts,
        })

    # Beste Vereine zuerst:
    # 1. Ø A absteigend
    # 2. Bester Tag A absteigend
    club_summary_rows.sort(
        key=lambda item: (
            item["average_a"] if item["average_a"] is not None else -1,
            item["best_a"] if item["best_a"] is not None else -1,
        ),
        reverse=True,
    )

    for rank, summary in enumerate(club_summary_rows, start=1):
        ws_clubs.cell(club_row, 1, summary["club_name"])

        if summary["cost"] is not None:
            ws_clubs.cell(club_row, 2, summary["cost"])

        ws_clubs.cell(club_row, 3, summary["games_count"])

        if summary["average_a"] is not None:
            ws_clubs.cell(club_row, 4, summary["average_a"])

        if summary["best_a"] is not None:
            ws_clubs.cell(club_row, 5, summary["best_a"])

        for offset, threshold in enumerate(thresholds, start=6):
            ws_clubs.cell(
                club_row,
                offset,
                summary["threshold_counts"][threshold],
            )

        for col in range(1, len(club_headers) + 1):
            style_body_cell(
                ws_clubs.cell(club_row, col),
                centered=(col != 1),
                bold=(col == 1),
            )

        # Top 3 der Vereinsübersicht dezent hervorheben.
        if rank <= 3:
            for col in range(1, len(club_headers) + 1):
                cell = ws_clubs.cell(club_row, col)
                cell.font = Font(
                    color=COLOR_GOLD,
                    bold=True,
                )

        ws_clubs.cell(
            club_row,
            2,
        ).number_format = '#,##0.00 [$€-de-DE]'

        ws_clubs.cell(
            club_row,
            4,
        ).number_format = '0.00'

        ws_clubs.cell(
            club_row,
            5,
        ).number_format = '0.00'

        club_row += 1

    last_club_col = "J" if len(club_headers) == 10 else "K"
    ws_clubs.auto_filter.ref = (
        f"A1:{last_club_col}{max(1, club_row - 1)}"
    )

    for col, width in {
        "A": 28, "B": 20, "C": 12, "D": 12, "E": 15,
        "F": 11, "G": 11, "H": 11, "I": 11, "J": 11, "K": 11,
    }.items():
        ws_clubs.column_dimensions[col].width = width

    game_headers = [
        "Verein", "Nr.", "Datum", "Gegner", "H/A",
        "TW", "VER", "MID", "FW", "Feldspieler",
        "Summe roh", "Summe A", "Marke A",
    ]
    ws_games.append(game_headers)
    style_header(ws_games, 1, 1, 13)
    ws_games.freeze_panes = "A2"
    game_row = 2

    for item in matchday_rows:
        values = [
            item["club_name"],
            item["number"],
            item["date"],
            item["opponent"],
            item["ha"],
            round(item["gk"], 2),
            round(item["def"], 2),
            round(item["mid"], 2),
            round(item["fwd"], 2),
            round(item["field"], 2),
            round(item["raw_sum"], 2),
            round(item["sum_a"], 2),
            item["mark_a"],
        ]

        for col, value in enumerate(values, start=1):
            ws_games.cell(game_row, col, value)
            style_body_cell(
                ws_games.cell(game_row, col),
                centered=(col not in (1, 4)),
                bold=(col == 1),
            )

        ws_games.cell(game_row, 3).number_format = "DD.MM.YYYY"

        for col in range(6, 13):
            ws_games.cell(game_row, col).number_format = "0.00"

        game_row += 1

    if matchday_rows:
        ws_games.auto_filter.ref = f"A1:M{game_row - 1}"

    for col, width in {
        "A": 28, "B": 7, "C": 13, "D": 28, "E": 8,
        "F": 10, "G": 10, "H": 10, "I": 10, "J": 13,
        "K": 13, "L": 13, "M": 12,
    }.items():
        ws_games.column_dimensions[col].width = width

    # --------------------------------------------------------
    # Seite 4: Spieltagsdetails
    # --------------------------------------------------------
    # Die bisherigen 3 Seiten bleiben unverändert.
    # Diese Zusatzseite zeigt zur Kontrolle, welche konkreten Spieler
    # pro Spieltag tatsächlich in die 5er-Wertung gekommen sind.

    detail_headers = [
        "Verein",
        "Nr.",
        "Datum",
        "Gegner",
        "H/A",
        "TW Spieler",
        "TW Score",
        "VER Spieler",
        "VER Score",
        "MID Spieler",
        "MID Score",
        "FW Spieler",
        "FW Score",
        "Feldspieler",
        "Feldspieler Score",
        "Kapitän",
        "Kapitän L40",
        "Summe roh",
        "Summe A",
        "Marke A",
    ]

    ws_details.append(detail_headers)
    style_header(ws_details, 1, 1, len(detail_headers))
    ws_details.row_dimensions[1].height = 34
    ws_details.freeze_panes = "A2"

    detail_row = 2

    for item in matchday_rows:
        values = [
            item["club_name"],
            item["number"],
            item["date"],
            item["opponent"],
            item["ha"],
            item["gk_name"],
            round(item["gk"], 2),
            item["def_name"],
            round(item["def"], 2),
            item["mid_name"],
            round(item["mid"], 2),
            item["fwd_name"],
            round(item["fwd"], 2),
            item["field_name"],
            round(item["field"], 2),
            item["captain_name"],
            round(item["captain_l40"], 2),
            round(item["raw_sum"], 2),
            round(item["sum_a"], 2),
            item["mark_a"],
        ]

        for col, value in enumerate(values, start=1):
            ws_details.cell(detail_row, col, value)
            style_body_cell(
                ws_details.cell(detail_row, col),
                centered=(col not in (1, 4, 6, 8, 10, 12, 14, 16)),
                bold=(col == 1),
            )

        ws_details.cell(detail_row, 3).number_format = "DD.MM.YYYY"

        for col in (7, 9, 11, 13, 15, 17, 18, 19):
            ws_details.cell(detail_row, col).number_format = "0.00"

        detail_row += 1

    if matchday_rows:
        ws_details.auto_filter.ref = f"A1:T{detail_row - 1}"

    detail_widths = {
        "A": 28,
        "B": 7,
        "C": 13,
        "D": 28,
        "E": 8,
        "F": 25,
        "G": 11,
        "H": 25,
        "I": 11,
        "J": 25,
        "K": 11,
        "L": 25,
        "M": 11,
        "N": 25,
        "O": 16,
        "P": 25,
        "Q": 13,
        "R": 13,
        "S": 13,
        "T": 11,
    }

    for col, width in detail_widths.items():
        ws_details.column_dimensions[col].width = width


    # --------------------------------------------------------
    # Seite 5: Meine Karten
    # --------------------------------------------------------
    my_card_headers = [
        "Sorare-Account",
        "Verein",
        "Position",
        "Spieler",
        "L40",
        "Startelf-Einsätze 26/27",
        "Seltenheit",
        "Saison",
        "Karten-Slug",
    ]

    ws_my_cards.append(my_card_headers)
    style_header(ws_my_cards, 1, 1, len(my_card_headers))
    ws_my_cards.freeze_panes = "A2"

    my_card_row = 2

    for card in owned_cards:
        values = [
            sorare_slug,
            card["club_name"],
            POSITION_LABELS[card["position"]],
            card["name"],
            round(card.get("l40", 0.0), 2),
            card.get("starts_2026_27", 0),
            card.get("rarity", "Limited"),
            "2026/27",
            card.get("card_slug", ""),
        ]

        for col, value in enumerate(values, start=1):
            ws_my_cards.cell(my_card_row, col, value)
            style_body_cell(
                ws_my_cards.cell(my_card_row, col),
                centered=(col in (3, 5, 6, 7, 8)),
                bold=(col == 4),
            )

        ws_my_cards.cell(my_card_row, 5).number_format = "0.00"
        my_card_row += 1

    if owned_cards:
        ws_my_cards.auto_filter.ref = f"A1:I{my_card_row - 1}"

    for col, width in {
        "A": 30,
        "B": 28,
        "C": 12,
        "D": 28,
        "E": 10,
        "F": 23,
        "G": 13,
        "H": 12,
        "I": 44,
    }.items():
        ws_my_cards.column_dimensions[col].width = width

    # --------------------------------------------------------
    # Seite 6: Mein bestes Lineup
    # --------------------------------------------------------
    best_headers = [
        "Spieltag",
        "Datum",
        "TW",
        "TW Score",
        "VER",
        "VER Score",
        "MID",
        "MID Score",
        "FW",
        "FW Score",
        "Feldspieler",
        "Feldspieler Score",
        "Kapitän",
        "Kapitän L40",
        "Summe roh",
        "Summe A",
        "Marke A",
        "5 Spieler vorhanden",
    ]

    ws_my_lineup.append(best_headers)
    style_header(ws_my_lineup, 1, 1, len(best_headers))
    ws_my_lineup.freeze_panes = "A2"

    lineup_row = 2

    for item in user_best_lineups:
        gk = item.get("gk")
        de = item.get("def")
        mi = item.get("mid")
        fw = item.get("fwd")
        extra = item.get("extra")
        captain = item.get("captain")

        values = [
            item["number"],
            item["date_text"],
            gk["name"] if gk else "-",
            round(gk["score"], 2) if gk else 0.0,
            de["name"] if de else "-",
            round(de["score"], 2) if de else 0.0,
            mi["name"] if mi else "-",
            round(mi["score"], 2) if mi else 0.0,
            fw["name"] if fw else "-",
            round(fw["score"], 2) if fw else 0.0,
            extra["name"] if extra else "-",
            round(extra["score"], 2) if extra else 0.0,
            captain["name"] if captain else "-",
            round(captain["l40"], 2) if captain else 0.0,
            round(item["raw_sum"], 2),
            round(item["sum_a"], 2),
            item["mark_a"],
            "JA" if item["complete"] else "NEIN",
        ]

        for col, value in enumerate(values, start=1):
            ws_my_lineup.cell(lineup_row, col, value)
            style_body_cell(
                ws_my_lineup.cell(lineup_row, col),
                centered=(col not in (3, 5, 7, 9, 11, 13)),
                bold=(col in (13, 16)),
            )

        for col in (4, 6, 8, 10, 12, 14, 15, 16):
            ws_my_lineup.cell(lineup_row, col).number_format = "0.00"

        lineup_row += 1

    if user_best_lineups:
        ws_my_lineup.auto_filter.ref = f"A1:R{lineup_row - 1}"

    for col, width in {
        "A": 10,
        "B": 25,
        "C": 27,
        "D": 11,
        "E": 27,
        "F": 11,
        "G": 27,
        "H": 11,
        "I": 27,
        "J": 11,
        "K": 27,
        "L": 16,
        "M": 27,
        "N": 13,
        "O": 13,
        "P": 13,
        "Q": 11,
        "R": 18,
    }.items():
        ws_my_lineup.column_dimensions[col].width = width

    safe_league = (
        league_name.replace("/", "-").replace("\\", "-").replace(" ", "_")
    )
    safe_user = "".join(
        char if char.isalnum() or char in ("-", "_") else "_"
        for char in sorare_slug
    )
    filename = (
        f"Sorare_Analyzer_{safe_league}_{safe_user}_"
        f"2026-27.xlsx"
    )
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(output_path)
    return output_path



# ============================================================
# DISCORD BOT
# ============================================================

intents = discord.Intents.default()
bot = commands.Bot(command_prefix="!", intents=intents)


LEAGUE_CHOICES = [
    app_commands.Choice(name="Bundesliga", value="bundesliga-de"),
    app_commands.Choice(name="Premier League", value="premier-league-gb-eng"),
    app_commands.Choice(name="La Liga", value="laliga-es"),
    app_commands.Choice(name="Ligue 1", value="ligue-1-fr"),
    app_commands.Choice(name="MLS", value="mlspa"),
    app_commands.Choice(name="Primeira Liga", value="primeira-liga-pt"),
    app_commands.Choice(name="Jupiler Pro League", value="jupiler-pro-league"),
    app_commands.Choice(
        name="Contender (2. BL + Ligue 2 + Österreich + HNL)",
        value=CONTENDER_SLUG,
    ),
]


def league_name_from_slug(competition_slug: str) -> str:
    if competition_slug == CONTENDER_SLUG:
        return "Contender"

    for name, slug in LEAGUES.values():
        if slug == competition_slug:
            return name
    return competition_slug


def run_full_analysis(
    discord_user_id: int,
    competition_slug: str,
) -> Tuple[str, dict]:
    """
    Kompletter Analyzer in EINEM Ablauf:
    1. Liga laden
    2. bisherige Vereinsanalyse
    3. Discord-ID -> Sorare-Account
    4. eigene In-Season Limited 2026/27 Karten laden
    5. persönliche Best-Lineups berechnen
    6. alles in EINE Excel-Datei schreiben
    """
    sorare_slug = DISCORD_SORARE_ACCOUNTS.get(discord_user_id)

    if not sorare_slug:
        raise RuntimeError(
            "Für deine Discord-ID ist noch kein Sorare-Account "
            "hinterlegt."
        )

    league_name = league_name_from_slug(competition_slug)

    print()
    print("=" * 80)
    print(f"DISCORD USER: {discord_user_id}")
    print(f"SORARE USER:  {sorare_slug}")
    print(f"LIGA:         {league_name}")
    print("=" * 80)

    competition_teams, team_competition = (
        get_combined_competition_teams(
            competition_slug
        )
    )

    if not competition_teams:
        raise RuntimeError("Keine Teams gefunden.")

    players = get_combined_league_players(
        competition_slug,
        competition_teams,
        team_competition,
    )

    if not players:
        raise RuntimeError("Keine gültigen Spieler gefunden.")

    add_l40(players)
    add_combined_starter_stats(
        players,
        competition_slug,
    )

    # Bisherige Liga-/Collection-Analyse bleibt bestehen.
    selected = build_selection(players)
    add_prices(selected)

    matchday_rows = build_matchday_rows(
        selected,
        competition_slug,
    )

    # Persönlicher Teil.
    all_owned_cards = get_user_inseason_limited_cards(
        sorare_slug
    )

    owned_cards, owned_unique_players = (
        match_owned_cards_to_league(
            all_owned_cards,
            players,
        )
    )

    user_best_lineups = build_user_best_lineups(
        owned_unique_players,
        competition_slug,
    )

    excel_path = create_excel(
        selected=selected,
        league_name=league_name,
        competition_slug=competition_slug,
        matchday_rows=matchday_rows,
        sorare_slug=sorare_slug,
        owned_cards=owned_cards,
        user_best_lineups=user_best_lineups,
    )

    info = {
        "sorare_slug": sorare_slug,
        "league_name": league_name,
        "owned_cards": len(owned_cards),
        "owned_players": len(owned_unique_players),
        "lineups": len(user_best_lineups),
    }

    return excel_path, info


@bot.tree.command(
    name="analyse",
    description="Erstellt deine komplette Sorare-Analyse als Excel-Datei.",
)
@app_commands.describe(
    liga="Welche Liga möchtest du analysieren?"
)
@app_commands.choices(liga=LEAGUE_CHOICES)
async def analyse(
    interaction: discord.Interaction,
    liga: app_commands.Choice[str],
):
    # Die Analyse kann deutlich länger als 15 Minuten dauern.
    # Discord-Interaction-Webhooks laufen nach ungefähr 15 Minuten ab.
    # Deshalb bestätigen wir den Slash-Befehl sofort und senden das fertige
    # Ergebnis später als normale Bot-Nachricht in denselben Kanal.
    await interaction.response.defer(
        thinking=True,
        ephemeral=False,
    )

    discord_user_id = interaction.user.id
    channel = interaction.channel

    if discord_user_id not in DISCORD_SORARE_ACCOUNTS:
        await interaction.edit_original_response(
            content=(
                "❌ Für deine Discord-ID ist noch kein Sorare-Account "
                "hinterlegt."
            )
        )
        return

    sorare_slug = DISCORD_SORARE_ACCOUNTS[discord_user_id]
    league_name = league_name_from_slug(liga.value)

    await interaction.edit_original_response(
        content=(
            f"⏳ **Analyse gestartet**\n"
            f"👤 Sorare: `{sorare_slug}`\n"
            f"🏆 Liga: **{league_name}**\n\n"
            f"Die Auswertung läuft. Ich sende die fertige Excel-Datei "
            f"hier in den Kanal, sobald sie fertig ist."
        )
    )

    excel_path = None

    try:
        excel_path, info = await asyncio.to_thread(
            run_full_analysis,
            discord_user_id,
            liga.value,
        )

        message = (
            f"✅ **Sorare Analyse fertig**\n"
            f"👤 Sorare: `{info['sorare_slug']}`\n"
            f"🏆 Liga: **{info['league_name']}**\n"
            f"🟡 Eigene passende Karten: **{info['owned_cards']}**\n"
            f"👥 Eindeutige eigene Spieler: **{info['owned_players']}**\n"
            f"📊 Persönliche Game Weeks: **{info['lineups']}**\n\n"
            f"Die Excel-Datei enthält die normale Liga-/Vereinsanalyse "
            f"sowie **Meine Karten** und **Mein bestes Lineup**."
        )

        if channel is None:
            raise RuntimeError(
                "Discord-Kanal konnte nicht ermittelt werden."
            )

        # WICHTIG:
        # Normale Kanal-Nachricht statt interaction.followup.send().
        # Dadurch funktioniert die Ausgabe auch dann noch, wenn die
        # Analyse länger als die Lebensdauer des Interaction-Tokens dauert.
        await channel.send(
            content=message,
            file=discord.File(excel_path),
        )

    except Exception as exc:
        error_message = (
            f"❌ **Analyse fehlgeschlagen**\n"
            f"`{exc}`"
        )

        if channel is not None:
            try:
                await channel.send(error_message)
            except Exception:
                print(error_message)
        else:
            print(error_message)

    finally:
        if excel_path and os.path.exists(excel_path):
            try:
                os.remove(excel_path)
            except OSError:
                pass


@bot.event
async def on_ready():
    print(f"Bot eingeloggt als: {bot.user}")
    print("Version: FINAL + MLS gesamte Saison 2026 inkl. Spieltage/Scores")

    try:
        if GUILD_ID:
            guild = discord.Object(id=GUILD_ID)
            bot.tree.copy_global_to(guild=guild)
            synced = await bot.tree.sync(guild=guild)
            print(
                f"{len(synced)} Befehle auf dem Server synchronisiert."
            )
        else:
            synced = await bot.tree.sync()
            print(
                f"{len(synced)} globale Befehle synchronisiert."
            )
    except Exception as exc:
        print(f"Fehler beim Synchronisieren: {exc}")


if __name__ == "__main__":
    bot.run(DISCORD_TOKEN)
